import os

from openpyxl import Workbook, load_workbook

import constant
import storage


SHEET_NAME = "dataset"


def build_row(movie: dict) -> list:
    row = []
    for feature in constant.MAIN_INFO:
        row.append(movie["main_info"][feature])
    for feature in constant.RAW_SCORES:
        row.append(movie["raw_scores"][feature])
    for feature in constant.TAGS_VIBE:
        row.append(movie[constant.TAGS_VIBE_SECTION][feature])
    return row


def export_dataset_to_excel(overwrite: bool = False) -> bool:
    data = storage.load_dataset()
    meta = storage.load_meta()
    os.makedirs(constant.DIR_TXT, exist_ok=True)

    if overwrite is False and os.path.exists(constant.EDIT_EXCEL):
        print(f'Excel для редактирования уже существует: {constant.EDIT_EXCEL}')
        print('Открываю существующий файл без перезаписи.')
        return True

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(constant.CSV_FIELDS)

    for movie in data.values():
        worksheet.append(build_row(movie))

    if len(data) == 0 and len(meta) > 0:
        for movie in meta.values():
            row = []
            for feature in constant.MAIN_INFO:
                row.append(movie["main_info"][feature])
            for feature in constant.RAW_SCORES:
                row.append(movie["raw_scores"][feature])
            row.extend([""] * len(constant.TAGS_VIBE))
            worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    try:
        workbook.save(constant.EDIT_EXCEL)
    except PermissionError:
        print(f'Не удалось открыть Excel для записи: {constant.EDIT_EXCEL}')
        print('Закрой файл в Excel и попробуй снова.')
        return False

    print(f'Excel для редактирования сохранен: {constant.EDIT_EXCEL}')
    return True


def load_movies_from_excel() -> list:
    try:
        workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    except PermissionError:
        print(f'Не удалось прочитать Excel: {constant.EDIT_EXCEL}')
        print('Закрой файл в Excel и попробуй снова.')
        return None

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)

        if header is None:
            print('Excel-файл пуст!')
            return None

        fieldnames = [str(value).strip() if value is not None else "" for value in header]
        if storage.is_supported_csv_fields(fieldnames) is False:
            print('Ошибка Excel! Заголовки не совпадают с ожидаемыми')
            print('Ожидались:', constant.CSV_FIELDS)
            print('Получены:', fieldnames)
            return None

        movies = []
        for row_number, values in enumerate(rows, start=2):
            if all(value is None or str(value).strip() == "" for value in values):
                continue

            row = {
                field: "" if value is None else str(value)
                for field, value in zip(fieldnames, values)
            }
            movie = storage.build_movie_from_row(row, row_number)
            if movie is None:
                return None
            movies.append(movie)

        if len(movies) == 0:
            print('В Excel нет записей для импорта.')
            return None

        return movies
    finally:
        workbook.close()


def replace_dataset_from_excel() -> bool:
    if os.path.exists(constant.EDIT_EXCEL) is False:
        print(f'Файл для редактирования не найден: {constant.EDIT_EXCEL}')
        return False

    movies = load_movies_from_excel()
    if movies is None:
        return False

    old_dataset = storage.load_dataset()
    old_meta = storage.load_meta()

    for file_name in [constant.FILE_NAME, constant.META_JSON]:
        if storage.is_file_writable(file_name) is False:
            print(f'Не удалось открыть файл для записи: {file_name}')
            print('Закрой файл в другой программе или проверь права доступа и попробуй снова.')
            return False

    storage.create_backup()
    try:
        storage.clean_dataset()
        storage.clean_meta()

        for movie in movies:
            if storage.add_movie(movie) is False:
                raise ValueError('Не удалось добавить запись из Excel.')
    except (PermissionError, ValueError) as error:
        storage.save_dataset(old_dataset)
        storage.save_meta(old_meta)
        print(f'Импорт Excel отменен: {error}')
        print('Старый dataset и meta восстановлены.')
        return False

    storage.rework_formated_scores()
    print(f'Dataset пересобран из Excel. Добавлено записей: {len(movies)}')
    return True
